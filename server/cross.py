import openpyxl
import mysql.connector
import pandas as pd

def insertData(dd1):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="cosmetic"
    )
    mycursor = mydb.cursor()

    sql = "INSERT INTO chemical ( cas , cname , cmname , per , st ,  img , des ,  bodypart , color ) VALUES (%s, %s , %s , %s , %s , %s , %s, %s, %s)"
    val =  (dd1[1] , dd1[0] , dd1[4] , dd1[3] ,"5" ,"-" ,"-" , dd1[2] , "-" )
    mycursor.execute(sql, val)
    mydb.commit()
    print(mycursor.rowcount, "record inserted.")


def getdata1() : 
    dataframe = openpyxl.load_workbook("C:/Users/praph/Downloads/cosmetic-products-regulation--annex-v---allowed-preservatives-export.xlsx") 
    dataframe1 = dataframe.active
    data = []
    i = 0
    for row in dataframe1.iter_rows(values_only=True) : 
        if i > 4 :
            row = list(row)
            if '%' in row[10]:
                row[10] = row[10][:-2]
            dd = [row[0], row[2], row[7], row[10]]
            data.append(dd)
        i += 1
    return data

def getdata2() :
    dataframe = openpyxl.load_workbook("C:/Users/praph/Downloads/v2.xlsx") 
    dataframe1 = dataframe.active
    data = []
    i = 0
    for row in dataframe1.iter_rows(values_only=True) : 
        if i > 7 :
            row = list(row)
            # if '%' in row[10]:
            #     row[10] = row[10][:-2]
            dd = [row[2], row[3]]
            data.append(dd)
        i += 1
    return data

def getCmname(cas, dd2):
    for x in dd2:
        if cas in str(x[1]):
            return x[0]
    return '-'

dd1 = getdata1()  # cname, cas, bodypart, per
dd2 = getdata2()  # cmname, cas
print(len(dd1))
count = 0
for i in range(len(dd1)):
    if dd1[i][1] != "-" : 
        w = getCmname(dd1[i][1], dd2)
        if w == None :
            w = "-"
        dd1[i].append(w)
        insertData( dd1[i] )
    # print(dd1[i])
    if w == '-':
        count += 1
print(count)


